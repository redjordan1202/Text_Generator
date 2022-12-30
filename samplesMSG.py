from openpyxl import Workbook,load_workbook
from datetime import datetime
from tkinter import *
from tkinter import filedialog, scrolledtext
from parsedatetime import Calendar
from time import sleep
import threading
import os
import re
import subprocess
import requests


class WorkOrder():
    order_number = None
    phone_number = None
    customer_name = None
    delivery_day = None
    start_time = None
    end_time = None
    address = None
    message = None

    def __str__(self):
        return self.order_number

message =  """Artificial Grass Delivery Confirmation- Your order, {}, has been dispatched and will be delivered {} between {} - {} at {}.To prepare for your delivery please make sure nothing is blocking the delivery location selected. You will receive another text notification 30 minutes prior to arrival. If there is a gate or entry approval, please provide and confirm."""

class MainWindow(Frame):
    def __init__(self, master = None):
        Frame.__init__(self, master)

        #Initial Window Set Up
        self.master = master
        master.title("Delivery Message Generator")
        self.master.resizable(0,0)                  #Setting window to not be resizable

        #Class Variables
        self.wb = None                              #Defining Workbook var. Keeping as None for now
        self.ws = None                              #Defining Worksheet var. Keeping as None for now
        self.path = None                            #Var to hold path to text file. Set to be same directory as the spreadsheet
        self.order_list = []                        #List of processed orders

        #Widget Definitions
        self.lbl_file = Label(master = self.master, text = "Selected File:")
        self.ent_file = Entry(
            master = self.master,
            state = "disabled",
            width = 75,
            justify = LEFT,
            disabledbackground = 'white',           #Overriding the default disabled background
            disabledforeground = 'black'            #Overriding the default disabled text color
        )
        self.btn_file = Button(
            master = self.master,
            text = "Choose File...",
            command = self.choose_file,
            width = 20
        )
        self.btn_run = Button(
            master = self.master,
            text = "Run",
            command = lambda: threading.Thread(target = self.process).start(),  #Start thread for processing
            width = 20
        )

        #Widget Deployment
        self.lbl_file.grid(column = 1, row = 0)
        self.ent_file.grid(column = 0, row = 1, columnspan = 3, pady = 10)
        self.btn_file.grid(column = 0, row = 2, padx = 50, pady = 10)
        self.btn_run.grid(column = 2, row = 2, padx = 50, pady = 10)

    #Class Functions

    #Edit File Entry
    def edit_entry(self,text):  
        self.ent_file.config(state = 'normal')          #Enabling the entry so I can be written
        self.ent_file.delete(0, END)                    #Deleting all text currently in the entry
        self.ent_file.insert(0,text)                    #Writing the new text to the entry

    #Create Check Messages Window
    def create_check_window(self):
        self.check_window = CheckWindow(self)

    #Used to write the order and its message to the text document
    def write_to_text(self, file, order):                 #Takes file path to txt document and current order object
        msg = """Customer Name: {}
Customer Number: {}
{}
================================================================================
"""
        file.write(msg.format(
            order.customer_name, 
            order.phone_number, 
            order.message
            )
        )

    #Function to handle spreadsheet opening
    def choose_file(self):  
        filetypes = (("Excel Spreadsheet", "*.xlsx"),)      #Defining filetypes to show in file dialogue 
        self.path = filedialog.askopenfilename(
            title = "Open File...",
            initialdir = os.path.expanduser('~'),           #Default directory is the Users "Home" Directory
            filetypes = filetypes
        )
        try:                                                #Attempt to load the workbook
            self.wb = load_workbook(filename= self.path)
            self.ws = self.wb.active
        except:                                             #Throw error if unable to open
            self.edit_entry("Make sure Excel file isn't open")
            return
        self.edit_entry(os.path.split(self.path)[1])        #Writing just the file name to the entry

    #Main Processing function
    def process(self):
        if self.path == None:                               #Check to ensure a file has been selected
            self.edit_entry("Please Select an Excel File")  #If not write error in file entry and return
            return

        self.edit_entry("Processing... Please Wait")
    
        txt_path = os.path.split(self.path)[0] + "/" + datetime.now().strftime("%m-%d-%Y") + ".txt" #Creating path to text in same directory as spreadsheet
        f = open(txt_path, "a")

        row = 1                                             #Set Initial Row number
        blank_rows = 0                                      #Var to count number of blank rows
        while row <= 10000:
            value = str(self.ws[("B" + str(row))].value)    #Pull value stored in Column be of current row
            if value.isnumeric() == True:                   #Check if value pulled is a number. If it is, its probably an order number
                order = WorkOrder()                         #Created new instance of WorkOrder
                blank_rows = 0                              #Reset blank row count as order was found
                order.order_number = value
                order.phone_number = re.sub('\D','',str(self.ws[("I" + str(row))].value))
                order.customer_name = str(self.ws[("D" + str(row))].value)
                self.get_time(row, order)                   #Send order to get_time function
                order.address = str(self.ws[("F" + str(row))].value)
                order.message = message.format(
                    order.order_number,
                    order.delivery_day, 
                    order.start_time,
                    order.end_time,
                    order.address,
                )
                self.order_list.append(order)               #Add current order to list of orders
                self.write_to_text(f, order)                #Write the order and message to file
                
            else:
                if value == "None":                         #Check if value is empty 
                    blank_rows = blank_rows + 1             #If it, increment blank_rows
                else:
                    blank_rows = 0                          #Reset blank row count as non-blank row was found
            
            if blank_rows > 4:                              #if we have 4 or more blank rows in a row
                break                                       #End the loop
            else:
                row = row + 1                               #Otherwise move to the next row

        self.edit_entry("Done!")
        f.close()                                           #Close the file as all writing is done
        subprocess.Popen(["notepad.exe", txt_path])         #Open the written text file so the user can send messages
        self.create_check_window()

    #Gets time from spreadsheet and calculates range. Formats time with AM/PM.
    def get_time(self, row, order):                         #Takes current row number and current order object
        value = str(self.ws[("H" + str(row))].value)
        cal = Calendar()                                    #Create calendar object so we can parse time
        time = cal.parse(value)                             #Convert human readable time to timedate object
        start_time = time[0].tm_hour                        #Set start_time to hour found before
        if start_time <= 3:                                 #Deliveries before 4m are set to 4am start time
            start_time = 4
        order.start_time = start_time
        order.end_time = start_time + 2                     #End time for range is 2 hours from start time

        #Format the time with 0 minutes and AM/PM markings
        if order.start_time > 12:
            order.start_time = str(order.start_time - 12)
            order.start_time = order.start_time + ":00 PM"
        elif order.start_time == 12:
            order.start_time = str(order.start_time) + ":00 PM"
        else:
            order.start_time = str(order.start_time) + ":00 AM"

        if order.end_time > 12:
            order.end_time = str(order.end_time - 12)
            order.end_time = order.end_time + ":00 PM"
        elif order.end_time == 12:
            order.end_time = str(order.end_time) + ":00 PM"
        else:
            order.end_time = str(order.end_time) + ":00 AM"

        if datetime.today().weekday() == 4:                 #Check if sending day is Friday. If it is change the message to say Monday delivery day
            order.delivery_day = "Monday"
        else:
            order.delivery_day = "tomorrow"

class CheckWindow(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.check_window = Toplevel()                      #Set the window as a toplevel window
        self.check_window.title("Check Messages")
        self.check_window.resizable(0,0)
        self.check_window.grab_set()                        #Have Window grab focus

        self.order = None                                   #Class var to hold current order being worked with
        self.order_list = StringVar(value = parent.order_list)  #Create Stringvar with order list as values

        #Widget Definitions
        self.frm_order_list = Frame(
            self.check_window,
            relief = "ridge",
            borderwidth = 4
        )
        self.lbl_order_list = Label(
            self.frm_order_list,
            text = "Work Orders",
            font = ("Ariel", 18)
        )
        self.lbl_order_count = Label(
            self.frm_order_list,
            text = "{} Orders Found".format(len(parent.order_list))
        )
        self.scrollbar = Scrollbar(self.frm_order_list)
        self.lbox_orders = Listbox(
            self.frm_order_list,
            selectmode = SINGLE,
            height = 25,
            width = 22,
            listvariable = self.order_list,
            yscrollcommand = self.scrollbar.set
        )
        self.lbox_orders.config(exportselection = False)    #Prevent listbox selection from un-highlighting on focus change
        self.scrollbar.config(command = self.lbox_orders.yview)  #Bind scrollbar to listbox
        self.frm_order_info = Frame(
            self.check_window,
            relief = "ridge",
            borderwidth = 4
        )
        self.lbl_order_info = Label(
            self.frm_order_info,
            text = "Order Information",
            font = ("Ariel", 18)
        )
        self.frm_info = Frame(self.frm_order_info)
        self.lbl_order_number = Label(
            self.frm_info,
            text = "Order Number"
        )
        self.lbl_order_phone = Label(
            self.frm_info,
            text = "Phone Number"
        )
        self.ent_order_number = Entry(
            self.frm_info,
            state = "disabled",
            width = 25,
            justify = LEFT,
            disabledbackground = 'white',
            disabledforeground = 'black'
        )
        self.ent_order_phone = Entry(
            self.frm_info,
            state = "disabled",
            width = 25,
            justify = LEFT,
            disabledbackground = 'white',
            disabledforeground = 'black'
        )
        self.lbl_message = Label(
            self.frm_info,
            text = "Message"
        )
        self.txt_message = Text(
            self.frm_info,
            width = 70,
            height = 20,
            wrap = WORD
        )
        self.btn_update = Button(
            self.frm_info,
            text = "Update Message",
            command = self.update_message
        )
        self.btn_confirm = Button(
            self.check_window,
            text = "Send Messages",
            command = lambda: threading.Thread(target = self.send_messages, args = (parent,)).start()
        )

        self.frm_log = Frame(
            self.check_window,
            relief = "ridge",
            borderwidth = 4
        )
        self.log = scrolledtext.ScrolledText(
            self.frm_log,
            width = 70,
            height = 20,
            wrap = WORD,
            state = DISABLED,
        )
        self.btn_finish = Button(
            self.frm_log,
            text = "Finish",
            state = DISABLED,
            command = parent.quit                           #Quit Program when finish button is clicked
        )
        
        """
        TODO
        - Change toplevel frames geometry to grid rather then pack
        - Move Send button to the bottom right corner of the window outside the rest of the frames
            - Maybe add an image to the send button. Something like an envelope
        """
        #Place Widgets
        self.frm_order_list.grid(column = 0, row = 0, padx = 5, pady =5)
        self.lbl_order_list.pack(pady = 5, padx = 5)
        self.lbl_order_count.pack(padx = 5, fill = X)
        self.lbox_orders.pack(side = LEFT, pady = 5, padx = 5, fill = X)
        self.scrollbar.pack(side = RIGHT, fill = Y)
        self.frm_order_info.grid(column = 1, row = 0, padx = 5, pady =5)
        self.lbl_order_info.pack(pady = 5, padx = 5)
        self.frm_info.pack(pady = 5, padx = 5)
        self.frm_info.pack_propagate(0)                     #Allow the use of Grid geometry within the frame
        self.lbl_order_number.grid(column = 0, row = 0)
        self.lbl_order_phone.grid(column = 1, row = 0)
        self.ent_order_number.grid(column = 0, row = 1)
        self.ent_order_phone.grid(column = 1, row = 1)
        self.lbl_message.grid(column = 0, row = 2, columnspan = 2)
        self.txt_message.grid(column = 0, row = 3, columnspan = 2)
        self.btn_update.grid(column = 1, row = 4, padx = 5, pady = 5, sticky = E)
        self.btn_confirm.grid(column = 1, row = 2, padx = 5, pady =5, sticky = E)
        
        self.lbox_orders.selection_anchor(0)                #Setting listbox section to the first entry
        self.lbox_orders.bind("<Double-1>", lambda event: self.get_order_info(parent))  #Bind double click to get_order_info
        self.lbox_orders.bind("<Map>", lambda event: self.get_order_info(parent))       #Run get_order_info when the listbox loads

    #Class Methods
    
    #Function to edit entry widgets
    def edit_entry(self,entry, text):  
        entry.config(state = 'normal')
        entry.delete(0, END)
        entry.insert(0,text) 
        entry.config(state = 'disabled')                    #Disable the entry box after writing
    
    #Function to edit the message text box
    def edit_text(self, content, text_box):
        text_box.delete(1.0, "end")                         #Delete all content in the box
        text_box.insert(1.0, content)                       #Write content starting at line 1,character 0

    #Function to write to sending log
    def write_log(self,text):
        self.log.config(state = NORMAL)
        self.log.insert("end", text)
        self.log.see(END)
        self.log.config(state = DISABLED)

    #Function to get and display order info    
    def get_order_info(self, parent):
        selection = self.lbox_orders.get(ANCHOR)            #Get current selection from listbox
        
        for i in parent.order_list:
            if i.order_number == selection:                 #Search order list for order number
                self.order = i                              
        
        #Write order info to the correct boxes
        self.edit_entry(self.ent_order_number, self.order.order_number)
        self.edit_entry(self.ent_order_phone, self.order.phone_number)
        self.edit_text(self.order.message, self.txt_message)

    #Function to update the saved order message
    def update_message(self):
        if self.order == None:                              #Check if an order is selected. Should always be selected
            return                                          #Return if somehow not selected
        self.order.message = self.txt_message.get(1.0, 'end-1c')    #Replace order message to be what is currently in the message text box

    #Function to send message 
    def send_messages(self,parent):
        #Create the log box
        self.frm_log.place(relx = 0.5, rely = 0.5, anchor = CENTER) #Placing the frame on top of all the other widgets
        self.log.pack(pady = 10, padx = 10)
        self.btn_finish.pack()

        #Disable other buttons and message text box
        self.btn_confirm.config(state = DISABLED)
        self.btn_update.config(state = DISABLED)
        self.txt_message.config(state = DISABLED)

        self.check_window.protocol("WM_DELETE_WINDOW", disable_event) #Disable X button on window

        hasError = False                                    #Bool value to track if an HTTP error was received

        #Write header messages to log
        self.write_log("Getting Ready to Send Messages\n")
        self.write_log("***PLEASE DO NOT CLOSE THE PROGRAM***\n")
        self.write_log("Processing {} Orders\n\n".format(len(parent.order_list)))
        sleep(5)                                            #Wait for a few seconds so that the user can read the above messages

        for order in parent.order_list:                     #Loop through all orders in the order list
            self.write_log("Sending Work Order {}\n".format(order.order_number))
            payload = {                                     #Create payload to send to server
                "destination" : order.phone_number,
                "source" : "1234567890",
                "clientMessageId" : order.order_number,
                "text" : order.message,
            }
            #Response is currently using a test REST api for now
            response = requests.post("https://eoinm2i8j9t7pkn.m.pipedream.net", payload)
            if response.status_code == 200:                 #Check if post request was successful
                self.write_log("Message Sent to 8x8 Successfully\n\n")
            else:                                           #If error is received throw an error message in the log
                self.write_log("***ERROR Code {}***\n".format(response.status_code))
                self.write_log("Message Failed to Send\n")
                hasError = True                             #Mark that an error has been encountered
            self.write_log("====================\n\n")
        self.write_log("Sending Complete\n")
        if hasError:                                        #If error was encountered, inform user
            self.write_log("*Warning* At Least Message had an Error During Sending")
        
        self.btn_finish.config(state = NORMAL)              #Allow user to click finish button to close program

#Function to disable on-click events
def disable_event():
    pass

if __name__ == "__main__":
    root = Tk()
    main_app = MainWindow(root)
    root.mainloop()
