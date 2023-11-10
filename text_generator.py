from tkinter import filedialog as fd
from openpyxl import load_workbook
from parsedatetime import Calendar
from datetime import datetime
from time import sleep
import os
import re
import subprocess
import sys


class WorkOrder:
    order_number = None
    phone_number = None
    customer_name = None
    delivery_day = None
    start_time = None
    end_time = None
    address = None
    message = None


columns = {
    "order_number": "A",
    "phone_number": "A",
    "customer_name": "A",
    "address": "A",
    "datetime": "A",
}

text_colors = {
    "green": '\033[92m',
    "red": '\033[91m',
    "endc": '\033[0m'
}

HEADER_ROW = 1

MESSAGE = """Artificial Grass Delivery Confirmation- Your order, {}, has been dispatched and will be delivered {} between {} - {} at {}.
To prepare for your delivery please make sure nothing is blocking the delivery location selected. 
You will receive another text notification 30 minutes prior to arrival. If there is a gate or entry approval, please provide and confirm."""


def main():
    has_error = False
    print(f"{text_colors['green']}Turf Distributors ETA Text Message Generator")
    print("Written by Jordan Del Pilar")
    print(f"jordan.delpilar@turfdistributors.com{text_colors['endc']}")
    print("=" * 80)
    print("\n")

    while True:
        filetypes = (("Excel Spreadsheet", "*.xlsx"),)
        wb_path = fd.askopenfilename(
            title="Select File...",
            initialdir=os.path.expanduser("~"),
            filetypes=filetypes
        )
        try:
            wb = load_workbook(filename=wb_path)
            ws = wb.active
            txt_path = os.path.split(wb_path)[0] + "/" + datetime.now().strftime("%m-%d-%Y") + ".txt"
            break
        except:
            print(f"{text_colors['red']}Could not open excel sheet\nMake sure the sheet is not open and try again.\n{text_colors['endc']}")
            input("Press Enter to try again")

    i = 1
    while i < 10:
        value = ws["A" + str(i)].value
        if "Service Resource: Name" in value:
            HEADER_ROW = i
            break
        else:
            i = i+1
            continue
    get_columns(ws, HEADER_ROW)

    while True:
        choice = input("Press S to start processing\nOr press E to exit:    ")
        choice = choice.upper()
        match choice:
            case "S":
                break
            case "E":
                print("Goodbye!")
                sleep(1)
                sys.exit("User Quit")

    os.system("cls")
    print("Processing Work Orders")
    order = WorkOrder
    f = open(txt_path, "a")
    wo_processed = 0
    i = 1
    while i < 1000:
        order_number = str(ws[columns["order_number"]+str(i)].value)
        if order_number:
            if (len(order_number) == 8) and order_number.isnumeric():
                print("Processing Work Order %s" % order_number)
                order.order_number = order_number
                order.customer_name = ws[columns["customer_name"] + str(i)].value
                order.address = ws[columns["address"] + str(i)].value

                # Process and format phone number
                phone_number = re.sub('\D', '', str(ws[columns["phone_number"] + str(i)].value))
                if phone_number:
                    if phone_number[0] == 1:
                        phone_number = "+" + phone_number
                    else:
                        phone_number = "+1" + phone_number
                    if len(phone_number) > 12:
                        phone_number = phone_number[:12]
                order.phone_number = phone_number

                # Parse start time
                raw_time = ws[columns["datetime"] + str(i)].value
                if type(raw_time) == datetime:
                    try:
                        raw_time = raw_time.strftime("%m/%d/%Y %H:%M:%S %p")
                    except:
                        try:
                            raw_time = raw_time.strftime("%m/%d/%Y %H:%M %p")
                        except:
                            print(f"{text_colors['red']}ERROR{text_colors['endc']} - Failed to convert date to string")
                            has_error = True
                            continue
                try:
                    cal = Calendar()
                    time = cal.parse(raw_time)
                except TypeError:
                    print(f"{text_colors['red']}ERROR{text_colors['endc']} - Failed to parse date")
                    has_error = True
                    continue

                # Set start time and convert from 24hr time to 12hr time
                start_time = time[0].tm_hour
                if start_time <= 3:
                    start_time = 4
                end_time = start_time + 2

                if start_time > 12:
                    order.start_time = str(start_time - 12) + ":00PM"
                elif start_time == 12:
                    order.start_time = str(start_time) + ":00PM"
                else:
                    order.start_time = str(start_time) + ":00AM"

                if end_time > 12:
                    order.end_time = str(end_time - 12) + ":00PM"
                elif end_time == 12:
                    order.end_time = str(end_time) + ":00PM"
                else:
                    order.end_time = str(end_time) + ":00AM"

                # Check if tomorrow is a weekday or not
                if datetime.today().weekday() == 4:
                    order.delivery_day = "Monday"
                else:
                    order.delivery_day = "tomorrow"

                order.message = MESSAGE.format(
                    order.order_number,
                    order.delivery_day,
                    order.start_time,
                    order.end_time,
                    order.address
                )

                print("Writing to file\n")
                try:
                    f.write("""Customer Name: %s
Customer Phone Number: %s
WO Number: %s

%s

================================================================================\n""" % (
                    order.customer_name,
                    order.phone_number,
                    order.order_number,
                    order.message
                ))
                except:
                    print(f"{text_colors['red']}ERROR{text_colors['endc']} - Failed to write to file.")
                    has_error = True
                    continue
                wo_processed = wo_processed + 1

        i = i + 1

    f.close()
    try:
        subprocess.Popen(["notepad.exe", txt_path])
    except:
        print(f"{text_colors['red']}ERROR{text_colors['endc']} - Failed to open notepad. Please open the file directly")
        has_error = True

    print(f"{text_colors['green']}")
    print("=" * 80)
    print("Processing Finished!")
    if has_error:
        print(f"{text_colors['red']}\n!!! Warning !!!{text_colors['endc']}") 
        print("There were some errors encounted during processing. Some Orders may not have been processed.")
        print("Please verify which orders were processed and report any errors displayed to Jordan\n")
        print(f"{text_colors['green']}")
        print("=" * 80)
    print(f"{wo_processed} Work Orders Processed")
    input(f"{text_colors['endc']}Press Enter to exit    ")


def get_columns(ws, header_row):
    print("Reading Column Headers...")
    header_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    for i in header_columns:
        header = ws[str(i) + str(header_row)].value
        if header:
            header = header.rstrip()

        match header:
            case "Work Order Number":
                columns["order_number"] = i
                continue
            case "Account: Account Name":
                columns["customer_name"] = i
                continue
            case "Address":
                columns["address"] = i
                continue
            case "Scheduled Start":
                columns["datetime"] = i
                continue
            case "Contact Phone Number":
                columns["phone_number"] = i
                continue
            case _:
                continue
    print("Column Headers reading finished!\n")


if __name__ == "__main__":
    main()
